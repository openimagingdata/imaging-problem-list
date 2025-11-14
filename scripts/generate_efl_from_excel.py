#!/usr/bin/env python3
"""
Generate EFL (Exam Finding List) JSON files from Excel spreadsheet.

This script reads an Excel file containing imaging findings organized by exam,
and generates one EFL JSON file per exam following the schema format.

Usage:
    python generate_efl_from_excel.py <excel_file> <output_dir> [--patient-mrn MRN] [--patient-dob DOB]
"""

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook


def generate_efl_files(excel_path: str, output_dir: str, patient_mrn: str = "MRN0000001", patient_dob: str = "1961-01-01"):
    """
    Generate EFL JSON files from Excel data.

    Args:
        excel_path: Path to the Excel file
        output_dir: Directory to write EFL JSON files
        patient_mrn: Patient MRN identifier
        patient_dob: Patient date of birth (YYYY-MM-DD)
    """
    # Load workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws is not None, "No active worksheet found"

    # Get column indices from header
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
    col_map = {}
    for cell in header_row:
        col_map[cell.value] = cell.column

    required_cols = ['Exam Date', 'Exam Type', 'Exam Code', 'Finding', 'OIDM Finding Model Name', 'OIDM FMID', 'Presence OIFMA_ID', 'Present/Absent', 'Text']
    for col_name in required_cols:
        if col_name not in col_map:
            raise ValueError(f"Required column '{col_name}' not found in Excel file")

    # Group findings by exam
    exams = defaultdict(list)

    for row_idx in range(2, ws.max_row + 1):
        exam_date = ws.cell(row=row_idx, column=col_map['Exam Date']).value
        exam_type = ws.cell(row=row_idx, column=col_map['Exam Type']).value
        exam_code = ws.cell(row=row_idx, column=col_map['Exam Code']).value
        finding_slug = ws.cell(row=row_idx, column=col_map['Finding']).value
        finding_name = ws.cell(row=row_idx, column=col_map['OIDM Finding Model Name']).value
        oifm_id = ws.cell(row=row_idx, column=col_map['OIDM FMID']).value
        oifma_id = ws.cell(row=row_idx, column=col_map['Presence OIFMA_ID']).value
        presence = ws.cell(row=row_idx, column=col_map['Present/Absent']).value
        text = ws.cell(row=row_idx, column=col_map['Text']).value

        if not all([exam_date, exam_type, exam_code, finding_slug, oifm_id, oifma_id, presence]):
            print(f"Warning: Row {row_idx} missing required data, skipping")
            continue

        exam_key = (exam_date, exam_type, exam_code)
        exams[exam_key].append({
            'finding_slug': finding_slug,
            'finding_name': finding_name,
            'oifm_id': oifm_id,
            'oifma_id': oifma_id,
            'presence': presence,
            'text': text if text else ''
        })

    # Create output directory
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    print(f"Found {len(exams)} unique exams")
    print()

    # Generate EFL for each exam
    for (exam_date, exam_type, exam_code), findings in sorted(exams.items()):
        # Bump date by 15 years
        if isinstance(exam_date, datetime):
            bumped_date = exam_date.replace(year=exam_date.year + 15)
        else:
            # Parse string date and bump by 15 years
            parsed_date = datetime.strptime(str(exam_date), "%Y-%m-%d")
            bumped_date = parsed_date.replace(year=parsed_date.year + 15)

        exam_datetime = bumped_date.strftime("%Y-%m-%dT10:00:00Z")
        exam_date_str = bumped_date.strftime("%Y-%m-%d")
        date_compact = bumped_date.strftime("%Y%m%d")

        # Generate IDs
        report_id = str(uuid4())

        # Study ID: first 2 words of exam type, uppercased and underscored
        exam_words = exam_type.split()[:2]
        study_prefix = '_'.join(word.upper() for word in exam_words)
        study_id = f"{study_prefix}_{date_compact}"

        # Create filename based on study ID (lowercase)
        filename = f"{study_id.lower()}_efl.json"

        # Build EFL structure
        efl = {
            "$schema": "https://github.com/openimagingdata/imaging-problem-list/schema/exam-problem-list-schema.json",
            "diagnosticReportId": report_id,
            "patientInfo": {
                "patientIdentifier": patient_mrn,
                "patientDOB": patient_dob
            },
            "examInfo": {
                "studyIdentifier": study_id,
                "studyDateTime": exam_datetime,
                "studyLoincCode": exam_code,
                "studyDescription": exam_type
            },
            "findings": []
        }

        # Add findings - track counts per finding slug for numbering
        finding_counts = defaultdict(int)

        for finding in findings:
            # Get finding slug
            finding_slug = finding['finding_slug']

            # Determine presence value and suffix
            presence_lower = finding['presence'].lower()
            if presence_lower == 'present':
                value_suffix = '.1'
                value_desc = 'present'
                # Increment counter for present findings
                finding_counts[finding_slug] += 1
                count = finding_counts[finding_slug]
            elif presence_lower == 'absent':
                value_suffix = '.0'
                value_desc = 'absent'
                # Absent findings always use _0
                count = 0
            elif presence_lower in ('indeterminate', 'uncertain'):
                value_suffix = '.2'
                value_desc = 'indeterminate'
                # Indeterminate findings always use _2
                count = 2
            else:
                print(f"Warning: Unknown presence value '{finding['presence']}' for {finding['oifm_id']}, defaulting to indeterminate")
                value_suffix = '.2'
                value_desc = 'indeterminate'
                count = 2

            # Generate observation ID
            observation_id = f"{finding_slug}_{count}"

            finding_obj = {
                "observationId": observation_id,
                "findingCode": finding['oifm_id'],
                "findingDescription": finding['finding_name'],
                "attributes": [
                    {
                        "attributeCode": finding['oifma_id'],
                        "attributeDescription": "presence",
                        "attributeValueCode": f"{finding['oifma_id']}{value_suffix}",
                        "attributeValueDescription": value_desc
                    }
                ]
            }

            # Only add reportText if it exists and is non-empty
            if finding['text']:
                finding_obj["reportText"] = finding['text']

            efl['findings'].append(finding_obj)

        # Write JSON file
        output_file = output_path / filename
        with open(output_file, 'w') as f:
            json.dump(efl, f, indent=2)

        print(f"Created: {filename} ({len(findings)} findings)")

    print()
    print(f"Total files created: {len(exams)}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate EFL JSON files from Excel spreadsheet"
    )
    parser.add_argument(
        "excel_file",
        help="Path to the Excel file"
    )
    parser.add_argument(
        "output_dir",
        help="Directory to write EFL JSON files"
    )
    parser.add_argument(
        "--patient-mrn",
        default="MRN0000001",
        help="Patient MRN identifier (default: MRN0000001)"
    )
    parser.add_argument(
        "--patient-dob",
        default="1961-01-01",
        help="Patient date of birth in YYYY-MM-DD format (default: 1961-01-01, ~63 years old in 2024)"
    )

    args = parser.parse_args()

    # Validate input file
    if not Path(args.excel_file).exists():
        print(f"Error: File not found: {args.excel_file}", file=sys.stderr)
        sys.exit(1)

    generate_efl_files(args.excel_file, args.output_dir, args.patient_mrn, args.patient_dob)


if __name__ == "__main__":
    main()
